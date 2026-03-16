# mahle_tecalliance_batch_v12.py
# V13: Batch scrape (Page2->Page3) + export Excel, with improved OE (原厂零件号/OE号) extraction.
#
# Change vs V11:
# - OE extraction now targets the OE table specifically and reads ONLY the first column links/text,
#   without over-filtering numeric formats (keeps "6 805 453" etc.)
#
# Usage:
#   python mahle_tecalliance_batch_v12.py "Input.xlsx" "Output.xlsx"
#   or just: python mahle_tecalliance_batch_v12.py  (defaults Input.xlsx/Output.xlsx)

import os
import re
import sys
import time
import traceback
import urllib.parse
from typing import List, Dict, Callable, Optional, Any

import pandas as pd
from playwright.sync_api import sync_playwright

HOME_URL = "https://web.tecalliance.net/mahle-catalog/zh/home"
SEARCH_URL = "https://web.tecalliance.net/mahle-catalog/zh/parts/search"
GROUPS = "448"
LOG_FILE = "batch_log.txt"


def ensure_workdir_is_script_dir() -> str:
    """Make double-click execution reliable by forcing cwd to the script directory."""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if script_dir:
            os.chdir(script_dir)
        return script_dir
    except Exception:
        return os.getcwd()


def log(msg: str):
    stamp = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{stamp}] {msg}"
    print(line)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def build_search_url(q: str) -> str:
    return f"{SEARCH_URL}?query={urllib.parse.quote(q, safe='')}&groups={GROUPS}"


def is_detail_url(url: str) -> bool:
    return bool(re.search(r"/parts/\d+/.+/detail", url or ""))


def dismiss_privacy_banner(page):
    for name in ["Reject all", "Agree to all", "拒絕全部", "同意全部", "全部同意", "全部拒絕"]:
        try:
            page.get_by_role("button", name=re.compile(name, re.I)).click(timeout=1500)
            page.wait_for_timeout(250)
            return
        except Exception:
            pass


def wait_detail(page, start_url: str, timeout_ms: int = 12000) -> bool:
    if is_detail_url(page.url) and page.url != start_url:
        return True
    try:
        page.wait_for_url(re.compile(r".*/parts/.*/detail.*"), timeout=timeout_ms)
        return True
    except Exception:
        return is_detail_url(page.url) and page.url != start_url


def mouse_click_in_box(page, loc, mode: str = "center") -> bool:
    box = loc.bounding_box()
    if not box:
        return False
    if mode == "topleft":
        x = box["x"] + min(18, box["width"] * 0.15)
        y = box["y"] + min(18, box["height"] * 0.20)
    else:
        x = box["x"] + box["width"] / 2
        y = box["y"] + box["height"] / 2
    page.mouse.click(x, y, delay=25)
    return True


def find_first_result_row(page):
    if page.locator("text=未找到商品").count() > 0:
        raise RuntimeError("未找到商品")

    rows = page.locator("[role='row']")
    if rows.count() > 0:
        for i in range(min(rows.count(), 8)):
            r = rows.nth(i)
            try:
                txt = (r.inner_text(timeout=1500) or "").strip()
            except Exception:
                txt = ""
            if not txt:
                continue
            if "商品編號" in txt and "名稱" in txt:
                continue
            if re.search(r"\b[A-Z]{1,4}\s*\d{2,3}\s*\d{3,4}", txt):
                return r
        if rows.count() >= 2:
            return rows.nth(1)

    code_re = re.compile(r"\b[A-Z]{1,4}\s*\d{2,3}\s*\d{3,4}\s*[A-Z0-9]{0,3}\b")
    code_loc = page.locator("a, td, div, span").filter(has_text=code_re).first
    if code_loc.count() == 0:
        raise RuntimeError("找不到任何商品列（無法定位第一筆結果）")
    row = code_loc.locator("xpath=ancestor::*[(@role='row') or self::tr][1]")
    if row.count() == 0:
        row = code_loc.locator("xpath=ancestor::div[1]")
    return row


def goto_page3_by_clicking_col2(page):
    dismiss_privacy_banner(page)
    page.wait_for_timeout(800)
    start_url = page.url

    row = find_first_result_row(page)

    td2 = row.locator("xpath=.//td[2]").first
    if td2.count() == 0:
        cells = row.locator(":scope > td, :scope > div")
        if cells.count() >= 2:
            td2 = cells.nth(1)
        else:
            all_td = row.locator("td")
            if all_td.count() >= 2:
                td2 = all_td.nth(1)
            else:
                raise RuntimeError("無法取得第2欄（商品圖片欄）的 cell（td/div）")

    try:
        td2.scroll_into_view_if_needed(timeout=2000)
    except Exception:
        pass

    for mode in ["center", "topleft", "center"]:
        mouse_click_in_box(page, td2, mode=mode)
        page.wait_for_timeout(250)
        if wait_detail(page, start_url, timeout_ms=6000):
            return

    box = td2.bounding_box()
    if box:
        for dx, dy in [(10, 10), (20, 12), (12, 22), (28, 18), (35, 25)]:
            page.mouse.click(box["x"] + dx, box["y"] + dy, delay=20)
            page.wait_for_timeout(250)
            if wait_detail(page, start_url, timeout_ms=5000):
                return

    raise RuntimeError("點擊第2欄（圖片欄）後仍未導到第3頁")


def norm(s: str) -> str:
    s = (s or "").replace("\u00ad", "")  # soft hyphen
    s = re.sub(r"\s+", " ", s).strip()
    return s


def extract_article_no(page) -> str:
    u = page.url
    m = re.search(r"/parts/\d+/([^/]+)/detail", u)
    if m:
        seg = m.group(1)
        try:
            decoded = urllib.parse.unquote(urllib.parse.unquote(seg))
            decoded = decoded.replace("+", " ")
            decoded = re.sub(r"\s+", " ", decoded).strip()
            if decoded:
                return decoded
        except Exception:
            pass
    t = norm(page.title())
    m2 = re.search(r"\b[A-Z]{1,4}\s*\d{2,3}\s*\d{3,4}\s*[A-Z0-9]{0,3}\b", t)
    return m2.group(0) if m2 else ""


def extract_value_after_label(page, label_variants: List[str]) -> str:
    for lab in label_variants:
        loc = page.locator(f"xpath=//*[normalize-space()={repr(lab)}]").first
        if loc.count():
            val = loc.locator("xpath=following-sibling::*[1]").first
            if val.count():
                try:
                    return norm(val.inner_text(timeout=1500))
                except Exception:
                    pass
            row = loc.locator("xpath=ancestor::*[self::tr or self::div][1]").first
            if row.count():
                try:
                    txt = norm(row.inner_text(timeout=1500))
                    if lab in txt:
                        return norm(txt.split(lab, 1)[1])
                except Exception:
                    pass
    for lab in label_variants:
        loc = page.locator(f"xpath=//*[contains(normalize-space(), {repr(lab)})]").first
        if loc.count():
            row = loc.locator("xpath=ancestor::*[self::tr or self::div][1]").first
            if row.count():
                try:
                    txt = norm(row.inner_text(timeout=1500))
                    parts = txt.split(lab, 1)
                    if len(parts) == 2:
                        return norm(parts[1])
                except Exception:
                    pass
    return ""


def extract_oe_numbers(page) -> List[str]:
    """
    Robust OE extraction:
    1) Locate the OE section header '原厂零件号（OE 号）' (or close variants).
    2) From that header, find the nearest following table (or the first table inside the same section),
       then extract ONLY first column values (anchors preferred, else td text).
    """
    hdr = page.locator("text=/原.*零件.*OE/i").first
    if hdr.count() == 0:
        hdr = page.locator("text=/原厂|原廠/i").filter(has_text=re.compile("OE", re.I)).first

    table = None
    if hdr.count():
        # try the closest table after header line
        table = hdr.locator("xpath=following::table[1]").first
        if table.count() == 0:
            # maybe table is within same panel/section
            panel = hdr.locator("xpath=ancestor::*[self::div or self::section][1]").first
            if panel.count():
                table = panel.locator("xpath=.//table").first

    if not table or table.count() == 0:
        # fallback: any table that has '原厂零件号' text near it
        candidate = page.locator("xpath=//table[.//th[contains(.,'OE')]]").first
        table = candidate if candidate.count() else None

    nums: List[str] = []
    if table and table.count():
        # Prefer anchors in first column
        a1 = table.locator("xpath=.//tr/td[1]//a[normalize-space()]")
        for i in range(min(a1.count(), 300)):
            t = norm(a1.nth(i).inner_text(timeout=1500))
            if t:
                nums.append(t)

        if not nums:
            td1 = table.locator("xpath=.//tr/td[1]")
            for i in range(min(td1.count(), 300)):
                t = norm(td1.nth(i).inner_text(timeout=1500))
                if t and t.lower() not in ["原厂零件号（oe 号）", "原廠零件號（oe 號）"]:
                    nums.append(t)

    # De-dup preserving order
    seen = set()
    out = []
    for n in nums:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def scrape_detail(page) -> Dict[str, str]:
    dismiss_privacy_banner(page)
    page.wait_for_timeout(600)

    detail_url = page.url
    article_no = extract_article_no(page)
    usage = extract_value_after_label(page, ["使用代碼", "使用代码", "使用代號", "使用代号"])
    repl = extract_value_after_label(page, ["替代", "更换为", "更換為", "更换為", "更換爲"])
    oe_list = extract_oe_numbers(page)

    return {
        "Detail_URL": detail_url,
        "ArticleNo": article_no,
        "OE_Numbers": "; ".join(oe_list),
        "Usage_Codes": usage,
        "Replacement": repl,
    }


def load_inputs(input_xlsx: str) -> List[str]:
    df = pd.read_excel(input_xlsx, sheet_name="Sheet1")
    if "Search_No" not in df.columns:
        raise RuntimeError("Input.xlsx 的 Sheet1 必須有欄位：Search_No")
    return [str(x).strip() for x in df["Search_No"].tolist()]


def save_output(rows: List[Dict[str, str]], output_xlsx: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["Search_No", "Status", "ArticleNo", "OE_Numbers", "Usage_Codes", "Replacement", "Detail_URL"]
    ws.append(headers)

    link_font = Font(color="0563C1", underline="single")
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        url = r.get("Detail_URL", "")
        if url:
            cell = ws.cell(row=ws.max_row, column=headers.index("Detail_URL") + 1)
            cell.hyperlink = url
            cell.font = link_font

    widths = {"A": 18, "B": 44, "C": 18, "D": 70, "E": 48, "F": 24, "G": 90}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A2"

    wb.save(output_xlsx)


def run(
    input_xlsx: str,
    output_xlsx: str,
    headless: bool = False,
    progress_cb: Optional[Callable[[int, int, str, str], None]] = None,
    cancel_cb: Optional[Callable[[], bool]] = None,
) -> Dict[str, Any]:
    # reset log file
    try:
        with open(LOG_FILE, "w", encoding="utf-8") as f:
            f.write("")
    except Exception:
        pass

    log(f"Python: {sys.version}")
    log(f"Working dir: {os.getcwd()}")
    log(f"Input: {input_xlsx}")
    log(f"Output: {output_xlsx}")

    queries = load_inputs(input_xlsx)
    total = len(queries)
    log(f"Total queries: {total}")
    if progress_cb:
        progress_cb(0, total, "", "START")

    results: List[Dict[str, str]] = []

    cancelled = False
    with sync_playwright() as p:
        # Render/Docker 需要 no-sandbox；--disable-dev-shm-usage 避免 /dev/shm 不足
        browser = p.chromium.launch(
            headless=headless,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--disable-software-rasterizer",
            ],
        )
        ctx = browser.new_context(locale="zh-TW")
        page = ctx.new_page()

        page.goto(HOME_URL, wait_until="domcontentloaded", timeout=60000)
        dismiss_privacy_banner(page)

        done = 0
        for q in queries:
            if cancel_cb and cancel_cb():
                cancelled = True
                log("CANCELLED by user request")
                if progress_cb:
                    progress_cb(done, total, "", "CANCELLED")
                break

            if not q:
                results.append({"Search_No": "", "Status": "SKIP(empty)"})
                done += 1
                if progress_cb:
                    progress_cb(done, total, "", "SKIP(empty)")
                continue

            log(f"Start: {q}")
            if progress_cb:
                progress_cb(done, total, q, "RUNNING")
            try:
                page.goto(build_search_url(q), wait_until="domcontentloaded", timeout=60000)
                goto_page3_by_clicking_col2(page)
                data = scrape_detail(page)
                results.append({"Search_No": q, "Status": "OK", **data})
                log(f"OK: {q} -> {data.get('ArticleNo','')} | OE={len(data.get('OE_Numbers','').split(';')) if data.get('OE_Numbers') else 0}")
                done += 1
                if progress_cb:
                    progress_cb(done, total, q, "OK")
            except Exception as e:
                tb = traceback.format_exc()
                log(f"ERROR: {q} -> {e}")
                log(tb)

                safe = re.sub(r"[^0-9A-Za-z]+", "_", q)[:50]
                try:
                    page.screenshot(path=f"batch_error_{safe}.png", full_page=True)
                    with open(f"batch_error_{safe}.html", "w", encoding="utf-8") as f:
                        f.write(page.content())
                except Exception:
                    pass

                results.append({"Search_No": q, "Status": f"ERROR: {e}", "Detail_URL": "", "ArticleNo": "", "OE_Numbers": "", "Usage_Codes": "", "Replacement": ""})
                done += 1
                if progress_cb:
                    progress_cb(done, total, q, f"ERROR: {e}")

        ctx.close()
        browser.close()

    save_output(results, output_xlsx)
    if cancelled:
        log("DONE (CANCELLED)")
    else:
        log("DONE")
    ok_count = sum(1 for r in results if r.get("Status") == "OK")
    err_count = sum(1 for r in results if str(r.get("Status", "")).startswith("ERROR"))
    skip_count = sum(1 for r in results if r.get("Status") == "SKIP(empty)")
    return {"total": len(results), "ok": ok_count, "error": err_count, "skip": skip_count, "cancelled": cancelled}


def main():
    ensure_workdir_is_script_dir()

    input_xlsx = sys.argv[1] if len(sys.argv) >= 2 else "Input.xlsx"
    output_xlsx = sys.argv[2] if len(sys.argv) >= 3 else "Output.xlsx"

    if not os.path.exists(input_xlsx):
        log(f"找不到 {input_xlsx}")
        input_xlsx = input("請輸入 Input.xlsx 完整路徑: ").strip().strip('"')
        if not input_xlsx:
            print("未提供路徑，結束")
            return

    run(input_xlsx, output_xlsx, headless=False)

    # Keep console open when double-clicked
    try:
        input('按 Enter 結束...')
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
